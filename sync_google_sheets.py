from __future__ import annotations

import argparse
import math
import re
import sys
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any

import gspread
from openpyxl import load_workbook

try:
    import tomllib
except ModuleNotFoundError:  # pragma: no cover - fallback for Python < 3.11.
    import tomli as tomllib  # type: ignore[no-redef]


ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL_PATH = ROOT_DIR / "Base de datos GEA.xlsx"
DEFAULT_SECRETS_PATH = Path(__file__).resolve().parent / ".streamlit" / "secrets.toml"

INVENTORY_SHEET_NAME = "Inventario_PT"
DISPATCH_SHEET_NAME = "Despachos"

INVENTORY_HEADERS = [
    "Producto",
    "Presentacion",
    "Stock actual",
    "Stock minimo",
    "Estado",
    "Ultima actualizacion",
]

DISPATCH_HEADERS = [
    "Fecha",
    "Hora",
    "Numero despacho",
    "Cliente",
    "Producto",
    "Presentacion",
    "Cantidad",
    "Estado",
    "Responsable",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0

    if isinstance(value, int | float):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)

    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_key(*parts: Any) -> tuple[str, ...]:
    return tuple(clean_text(part).casefold() for part in parts)


def read_table(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_name}' en el Excel local.")

    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [clean_text(value) for value in rows[0]]
    records: list[dict[str, Any]] = []

    for raw_row in rows[1:]:
        record = {
            headers[index]: raw_row[index] if index < len(raw_row) else None
            for index in range(len(headers))
            if headers[index]
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)

    return records


def split_date_time(value: Any) -> tuple[str, str]:
    if value is None or value == "":
        return "", ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d"), value.strftime("%H:%M:%S")

    text = str(value).strip()
    for date_format in (
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y",
    ):
        try:
            parsed = datetime.strptime(text, date_format)
            return parsed.strftime("%Y-%m-%d"), parsed.strftime("%H:%M:%S")
        except ValueError:
            pass

    match = re.match(r"^(\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{4})\s+(.+)$", text)
    if match:
        return match.group(1), match.group(2)

    return text, ""


def build_inventory_rows(workbook: Any) -> list[list[Any]]:
    inventory_records = read_table(workbook, "Inventario PT")
    minimum_records = read_table(workbook, "Minimos PT")

    minimums = {
        normalize_key(row.get("Producto"), row.get("Presentacion")): to_number(
            row.get("Stock minimo unidades")
        )
        for row in minimum_records
        if clean_text(row.get("Producto")) and clean_text(row.get("Presentacion"))
    }

    stock_by_key: dict[tuple[str, ...], dict[str, Any]] = {}
    for row in inventory_records:
        product = clean_text(row.get("Producto"))
        presentation = clean_text(row.get("Presentacion"))
        if not product or not presentation:
            continue

        units = to_number(row.get("Stock en unidades"))
        if units <= 0:
            continue

        key = normalize_key(product, presentation)
        item = stock_by_key.setdefault(
            key,
            {
                "Producto": product,
                "Presentacion": presentation,
                "Stock actual": 0.0,
            },
        )
        item["Stock actual"] += units

    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    output_rows: list[list[Any]] = []

    for key in sorted(set(stock_by_key) | set(minimums), key=lambda item: (item[0], item[1])):
        item = stock_by_key.get(
            key,
            {
                "Producto": key[0].title(),
                "Presentacion": key[1].title(),
                "Stock actual": 0.0,
            },
        )
        stock = float(item["Stock actual"])
        minimum = float(minimums.get(key, 0.0))

        if stock <= 0:
            status = "Sin stock"
        elif minimum > 0 and stock < minimum:
            status = "Bajo stock"
        else:
            status = "OK"

        output_rows.append(
            [
                item["Producto"],
                item["Presentacion"],
                stock,
                minimum,
                status,
                updated_at,
            ]
        )

    return output_rows


def build_dispatch_rows(workbook: Any) -> list[list[Any]]:
    dispatch_records = read_table(workbook, "Despachos")
    detail_records = read_table(workbook, "Detalle Despacho")

    dispatch_by_code = {
        clean_text(row.get("Codigo despacho")).casefold(): row
        for row in dispatch_records
        if clean_text(row.get("Codigo despacho"))
    }

    output_rows: list[list[Any]] = []
    for detail in detail_records:
        dispatch_code = clean_text(detail.get("Codigo despacho"))
        if not dispatch_code:
            continue

        header = dispatch_by_code.get(dispatch_code.casefold(), {})
        date_text, time_text = split_date_time(header.get("Fecha despacho real"))

        output_rows.append(
            [
                date_text,
                time_text,
                dispatch_code,
                clean_text(header.get("Cliente")),
                clean_text(detail.get("Producto")),
                clean_text(detail.get("Presentacion")),
                to_number(detail.get("Unidades despachadas")),
                clean_text(header.get("Estado")),
                clean_text(header.get("Responsable")),
            ]
        )

    return output_rows


def load_gsheets_config(secrets_path: Path) -> dict[str, Any]:
    if not secrets_path.exists():
        raise FileNotFoundError(f"No existe el archivo de secretos: {secrets_path}")

    with secrets_path.open("rb") as file:
        secrets = tomllib.load(file)

    gsheets_config = secrets.get("connections", {}).get("gsheets", {})
    required_keys = [
        "spreadsheet",
        "type",
        "project_id",
        "private_key_id",
        "private_key",
        "client_email",
        "client_id",
        "auth_uri",
        "token_uri",
        "auth_provider_x509_cert_url",
        "client_x509_cert_url",
    ]
    missing = [key for key in required_keys if not clean_text(gsheets_config.get(key))]
    if missing:
        raise ValueError(f"Faltan datos en [connections.gsheets]: {', '.join(missing)}")

    return gsheets_config


def update_worksheet(
    spreadsheet: Any,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
) -> int:
    try:
        worksheet = spreadsheet.worksheet(title)
        worksheet.clear()
    except gspread.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(
            title=title,
            rows=max(len(rows) + 10, 50),
            cols=len(headers) + 5,
        )

    values = [headers] + rows
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        try:
            worksheet.update(values=values, range_name="A1", value_input_option="USER_ENTERED")
        except TypeError:
            worksheet.update("A1", values, value_input_option="USER_ENTERED")
    worksheet.freeze(rows=1)

    try:
        worksheet.format(
            "1:1",
            {
                "textFormat": {"bold": True},
                "backgroundColor": {"red": 0.9, "green": 0.93, "blue": 0.97},
            },
        )
    except Exception:
        # Formatting is cosmetic; data sync should not fail because of it.
        pass

    return len(rows)


def sync_google_sheets(excel_path: Path, secrets_path: Path, dry_run: bool = False) -> dict[str, int]:
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el Excel local: {excel_path}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    inventory_rows = build_inventory_rows(workbook)
    dispatch_rows = build_dispatch_rows(workbook)
    workbook.close()

    result = {
        INVENTORY_SHEET_NAME: len(inventory_rows),
        DISPATCH_SHEET_NAME: len(dispatch_rows),
    }

    if dry_run:
        return result

    gsheets_config = load_gsheets_config(secrets_path)
    spreadsheet_url = str(gsheets_config.pop("spreadsheet"))
    google_client = gspread.service_account_from_dict(gsheets_config)
    spreadsheet = google_client.open_by_url(spreadsheet_url)

    result[INVENTORY_SHEET_NAME] = update_worksheet(
        spreadsheet,
        INVENTORY_SHEET_NAME,
        INVENTORY_HEADERS,
        inventory_rows,
    )
    result[DISPATCH_SHEET_NAME] = update_worksheet(
        spreadsheet,
        DISPATCH_SHEET_NAME,
        DISPATCH_HEADERS,
        dispatch_rows,
    )

    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Sincroniza el Excel local de GEA hacia Google Sheets para el dashboard PT."
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help=f"Ruta del Excel maestro. Por defecto: {DEFAULT_EXCEL_PATH}",
    )
    parser.add_argument(
        "--secrets",
        type=Path,
        default=DEFAULT_SECRETS_PATH,
        help=f"Ruta de secrets.toml. Por defecto: {DEFAULT_SECRETS_PATH}",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Calcula los registros a sincronizar, pero no escribe en Google Sheets.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    try:
        result = sync_google_sheets(
            excel_path=args.excel.resolve(),
            secrets_path=args.secrets.resolve(),
            dry_run=args.dry_run,
        )
    except Exception as exc:  # noqa: BLE001 - CLI entrypoint should show concise operational failures.
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    action = "Verificados" if args.dry_run else "Sincronizados"
    print(f"{action} {result[INVENTORY_SHEET_NAME]} registros en {INVENTORY_SHEET_NAME}.")
    print(f"{action} {result[DISPATCH_SHEET_NAME]} registros en {DISPATCH_SHEET_NAME}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
